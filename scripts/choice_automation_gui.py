import ctypes
import ctypes.wintypes
import logging
import queue
import subprocess
import sys
import threading
import tempfile
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook

import choice_automation


def get_app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = get_app_dir()
SETTINGS_PATH = APP_DIR / "choice_automation_gui_settings.json"
ERROR_LOG_PATH = Path(tempfile.gettempdir()) / "choice_downloader_gui_error.log"
DEFAULT_BROWSE_DIR = Path.home() if Path.home().exists() else APP_DIR
KEYWORD_MATCH_MODE_LABELS = {
    "满足任一关键词": "or",
    "满足全部关键词": "and",
}
KEYWORD_MATCH_MODE_VALUES = list(KEYWORD_MATCH_MODE_LABELS.keys())
COORDINATE_TARGET_LABELS = [
    ("left_nav_scroll", "左侧导航滚动点"),
    ("company_announcement", "公司公告入口"),
    ("all_announcements", "全部公告筛选"),
    ("financial_report", "财务报告筛选"),
    ("batch_download", "批量下载按钮"),
    ("popup_browse", "弹窗浏览按钮"),
    ("popup_range_checkbox", "弹窗范围勾选"),
    ("popup_range_input", "弹窗篇数输入"),
    ("popup_download", "弹窗下载按钮"),
    ("return_home", "返回首页按钮"),
]
COORDINATE_LABEL_BY_NAME = dict(COORDINATE_TARGET_LABELS)


class QueueLogHandler(logging.Handler):
    def __init__(self, output_queue: queue.Queue[str]):
        super().__init__()
        self.output_queue = output_queue
        self.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

    def emit(self, record: logging.LogRecord):
        try:
            self.output_queue.put(self.format(record) + "\n")
        except Exception:
            pass


class AutomationGui:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Choice 公告批量下载工具")
        self.root.geometry("1560x1040")
        self.root.minsize(1360, 880)

        self.output_queue: queue.Queue = queue.Queue()
        self.worker_thread: threading.Thread | None = None
        self.calibration_thread: threading.Thread | None = None
        self.is_running = False
        self.is_calibrating = False
        self.stop_requested = False
        self.stop_event = threading.Event()
        self.log_handler = QueueLogHandler(self.output_queue)

        self.excel_path_var = tk.StringVar()
        self.choice_exe_var = tk.StringVar()
        self.output_dir_var = tk.StringVar()
        self.log_dir_var = tk.StringVar()
        self.batch_count_var = tk.StringVar()
        self.max_batch_count_var = tk.StringVar()
        self.range_start_var = tk.StringVar()
        self.range_end_var = tk.StringVar()
        self.filename_keywords_var = tk.StringVar()
        self.keyword_match_mode_var = tk.StringVar(value="满足任一关键词")
        self.latest_only_var = tk.BooleanVar(value=False)
        self.start_wait_var = tk.StringVar()
        self.enter_wait_var = tk.StringVar()
        self.post_f9_wait_var = tk.StringVar()
        self.navigation_only_var = tk.BooleanVar(value=False)
        self.skip_navigation_var = tk.BooleanVar(value=False)
        self.skip_folder_dialog_var = tk.BooleanVar(value=False)
        self.return_home_var = tk.BooleanVar(value=True)
        self.coordinate_vars: dict[str, tuple[tk.StringVar, tk.StringVar]] = {}
        self.confirmed_coordinates = dict(choice_automation.COORDINATE_DEFAULTS)
        for target_name, _label in COORDINATE_TARGET_LABELS:
            default_x, default_y = choice_automation.COORDINATE_DEFAULTS[target_name]
            self.coordinate_vars[target_name] = (tk.StringVar(value=str(default_x)), tk.StringVar(value=str(default_y)))
        self.calibration_target_var = tk.StringVar(value=COORDINATE_TARGET_LABELS[1][1])

        self.status_var = tk.StringVar(value="就绪")
        self.summary_var = tk.StringVar(value="请填写本机路径和运行参数。程序不会预置或保存用户目录、公司名单和筛选条件。")
        self.company_count_var = tk.StringVar(value="未选择 Excel 文件")
        self.excel_path_var.trace_add("write", self._on_excel_path_changed)

        self._load_settings()
        self._configure_style()
        self._build_ui()
        self._refresh_company_count_text()
        self._drain_output_queue()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _configure_style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("App.TFrame", background="#eef3f9")
        style.configure("Header.TFrame", background="#dfeaf8")
        style.configure("Panel.TFrame", background="#eef3f9")
        style.configure("Surface.TFrame", background="#f7fafd")
        style.configure("Card.TLabelframe", background="#f7fafd", borderwidth=1, relief="solid", padding=14)
        style.configure("Card.TLabelframe.Label", background="#f7fafd", font=("Microsoft YaHei UI", 10, "bold"), foreground="#17324d")
        style.configure("Title.TLabel", background="#dfeaf8", font=("Microsoft YaHei UI", 20, "bold"), foreground="#10233f")
        style.configure("Subtle.TLabel", foreground="#5f6b7a", background="#f7fafd")
        style.configure("Hero.TLabel", background="#dfeaf8", font=("Microsoft YaHei UI", 10), foreground="#29486b")
        style.configure("Status.TLabel", font=("Microsoft YaHei UI", 10, "bold"), foreground="#133152", background="#eef3f9")
        style.configure("Stat.TLabel", background="#dfeaf8", foreground="#17324d", font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("Accent.TButton", font=("Microsoft YaHei UI", 10, "bold"), padding=(10, 8))
        style.map("Accent.TButton", background=[("active", "#1f5d99"), ("!disabled", "#2d6fb1")], foreground=[("!disabled", "#ffffff")])
        style.configure("Muted.TButton", padding=(10, 8))
        style.configure("App.Vertical.TScrollbar", arrowsize=12)

    def _build_ui(self):
        self.root.configure(bg="#eef3f9")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        header = ttk.Frame(self.root, style="Header.TFrame", padding=(18, 10, 18, 8))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=3, uniform="header_cols")
        header.columnconfigure(1, weight=2, uniform="header_cols")
        ttk.Label(header, text="Choice 公告批量下载工具", style="Title.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(header, text="从用户提供的 Excel 公司名单批量下载公告文件", style="Hero.TLabel").grid(row=1, column=0, sticky="w", pady=(4, 0))
        ttk.Label(header, textvariable=self.summary_var, style="Hero.TLabel", wraplength=820, justify="left").grid(
            row=2, column=0, sticky="w", pady=(6, 0)
        )

        stat_panel = ttk.Frame(header, style="Header.TFrame")
        stat_panel.grid(row=0, column=1, rowspan=3, sticky="nsew", padx=(18, 0))
        stat_panel.columnconfigure(0, weight=1)
        stat_panel.columnconfigure(1, weight=1)
        ttk.Label(stat_panel, text="Excel 状态", style="Hero.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(stat_panel, textvariable=self.company_count_var, style="Stat.TLabel", wraplength=320, justify="left").grid(
            row=1, column=0, sticky="w", pady=(6, 0)
        )
        ttk.Label(stat_panel, text="当前状态", style="Hero.TLabel").grid(row=0, column=1, sticky="w", padx=(20, 0))
        ttk.Label(stat_panel, textvariable=self.status_var, style="Stat.TLabel").grid(row=1, column=1, sticky="w", padx=(20, 0), pady=(6, 0))

        paned = ttk.Panedwindow(self.root, orient="vertical")
        paned.grid(row=1, column=0, sticky="nsew", padx=16, pady=(10, 8))

        top_surface = ttk.Frame(paned, style="Surface.TFrame", padding=0)
        top_surface.columnconfigure(0, weight=1)
        top_surface.rowconfigure(0, weight=1)

        self.content_canvas = tk.Canvas(
            top_surface,
            background="#eef3f9",
            highlightthickness=0,
            bd=0,
        )
        content_scrollbar = ttk.Scrollbar(
            top_surface,
            orient="vertical",
            command=self.content_canvas.yview,
            style="App.Vertical.TScrollbar",
        )
        self.content_canvas.configure(yscrollcommand=content_scrollbar.set)
        self.content_canvas.grid(row=0, column=0, sticky="nsew")
        content_scrollbar.grid(row=0, column=1, sticky="ns")

        self.scrollable_body = ttk.Frame(self.content_canvas, style="Panel.TFrame", padding=(0, 0, 4, 0))
        self.scrollable_body.columnconfigure(0, weight=1, uniform="top_cols")
        self.scrollable_body.columnconfigure(1, weight=1, uniform="top_cols")
        self.content_canvas_window = self.content_canvas.create_window((0, 0), window=self.scrollable_body, anchor="nw")
        self.scrollable_body.bind("<Configure>", self._on_scroll_body_configure)
        self.content_canvas.bind("<Configure>", self._on_canvas_configure)

        left = ttk.Frame(self.scrollable_body, style="Panel.TFrame")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)

        right = ttk.Frame(self.scrollable_body, style="Panel.TFrame")
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        form = ttk.LabelFrame(left, text="任务配置（请按本机环境填写）", style="Card.TLabelframe")
        form.grid(row=0, column=0, sticky="nsew")
        form.columnconfigure(1, weight=1)

        self._add_path_row(form, 0, "Excel 文件", self.excel_path_var, self._choose_excel_file, "选择文件")
        ttk.Label(form, text="支持 .xlsx / .xlsm，第一行需要包含“公司名称”列。", style="Subtle.TLabel").grid(
            row=1, column=1, columnspan=2, sticky="w", padx=(10, 0), pady=(0, 8)
        )

        self._add_path_row(form, 2, "Choice 程序/快捷方式", self.choice_exe_var, self._choose_choice_exe, "系统浏览")
        self._add_path_row(form, 3, "下载目录", self.output_dir_var, self._choose_output_dir, "选择目录")
        self._add_path_row(form, 4, "日志目录", self.log_dir_var, self._choose_log_dir, "选择目录")

        ttk.Label(form, text="每家公司下载篇数").grid(row=5, column=0, sticky="w", pady=8)
        ttk.Spinbox(form, from_=1, to=999, textvariable=self.batch_count_var, width=10).grid(
            row=5, column=1, sticky="w", padx=(10, 0), pady=8
        )
        batch_limit_grid = ttk.Frame(form)
        batch_limit_grid.grid(row=5, column=2, sticky="w", pady=8)
        ttk.Label(batch_limit_grid, text="本次上限").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(batch_limit_grid, from_=1, to=999, textvariable=self.max_batch_count_var, width=8).grid(
            row=0, column=1, sticky="w", padx=(8, 0)
        )

        range_grid = ttk.Frame(form)
        range_grid.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        ttk.Label(range_grid, text="提取范围").grid(row=0, column=0, sticky="w")
        ttk.Label(range_grid, text="第").grid(row=0, column=1, sticky="w", padx=(8, 4))
        ttk.Spinbox(range_grid, from_=1, to=99999, textvariable=self.range_start_var, width=8).grid(row=0, column=2, sticky="w")
        ttk.Label(range_grid, text="家 到 第").grid(row=0, column=3, sticky="w", padx=(8, 4))
        ttk.Spinbox(range_grid, from_=1, to=99999, textvariable=self.range_end_var, width=8).grid(row=0, column=4, sticky="w")
        ttk.Label(range_grid, text="家").grid(row=0, column=5, sticky="w", padx=(4, 0))

        wait_grid = ttk.Frame(form)
        wait_grid.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        ttk.Label(wait_grid, text="启动等待").grid(row=0, column=0, sticky="w")
        ttk.Entry(wait_grid, textvariable=self.start_wait_var, width=8).grid(row=0, column=1, sticky="w", padx=(8, 18))
        ttk.Label(wait_grid, text="Enter 等待").grid(row=0, column=2, sticky="w")
        ttk.Entry(wait_grid, textvariable=self.enter_wait_var, width=8).grid(row=0, column=3, sticky="w", padx=(8, 18))
        ttk.Label(wait_grid, text="F9 等待").grid(row=0, column=4, sticky="w")
        ttk.Entry(wait_grid, textvariable=self.post_f9_wait_var, width=8).grid(row=0, column=5, sticky="w", padx=(8, 0))

        filter_grid = ttk.Frame(form)
        filter_grid.grid(row=8, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        filter_grid.columnconfigure(1, weight=1)
        ttk.Label(filter_grid, text="文件名筛选").grid(row=0, column=0, sticky="w")
        ttk.Entry(filter_grid, textvariable=self.filename_keywords_var).grid(row=0, column=1, sticky="ew", padx=(10, 0))
        ttk.Label(filter_grid, text="匹配方式").grid(row=1, column=0, sticky="w", pady=(8, 0))
        ttk.Combobox(
            filter_grid,
            textvariable=self.keyword_match_mode_var,
            values=KEYWORD_MATCH_MODE_VALUES,
            state="readonly",
            width=14,
        ).grid(row=1, column=1, sticky="w", padx=(10, 0), pady=(8, 0))
        ttk.Checkbutton(filter_grid, text="只保留最新日期", variable=self.latest_only_var).grid(
            row=2, column=1, sticky="w", pady=(8, 0)
        )

        ttk.Label(
            form,
            text="文件名筛选为可选项。可填写逗号分隔关键词，例如 10-K,20-F,40-F；留空表示保留全部下载文件。",
            style="Subtle.TLabel",
            wraplength=640,
            justify="left",
        ).grid(row=9, column=0, columnspan=3, sticky="w", pady=(12, 0))

        tips = ttk.LabelFrame(left, text="填写提示", style="Card.TLabelframe")
        tips.grid(row=1, column=0, sticky="nsew", pady=(10, 0))
        tips.columnconfigure(0, weight=1)
        ttk.Label(
            tips,
            text=(
                "1. 路径和运行参数均由用户自行填写，本软件不预置个人目录或筛选条件。\n"
                "2. Choice 入口可以选择安装目录下的 exe，也可以选择桌面快捷方式 lnk。\n"
                "3. 下载目录和日志目录会在运行前自动创建。\n"
                "4. 批量运行不要勾选“跳过导航”。\n"
                "5. 等待时间单位为秒；电脑或网络较慢时可适当调大。"
            ),
            justify="left",
            style="Subtle.TLabel",
            wraplength=620,
        ).grid(row=0, column=0, sticky="w")

        mode_card = ttk.LabelFrame(right, text="运行选项", style="Card.TLabelframe")
        mode_card.grid(row=0, column=0, sticky="nsew")
        mode_card.columnconfigure(0, weight=1)

        ttk.Checkbutton(mode_card, text="只导航到公告页，不执行下载", variable=self.navigation_only_var).grid(row=0, column=0, sticky="w", pady=6)
        ttk.Checkbutton(mode_card, text="跳过导航，假设当前已经在公告页", variable=self.skip_navigation_var).grid(row=1, column=0, sticky="w", pady=6)
        ttk.Checkbutton(mode_card, text="跳过下载弹窗里的目录选择步骤", variable=self.skip_folder_dialog_var).grid(row=2, column=0, sticky="w", pady=6)
        ttk.Checkbutton(mode_card, text="最后一家公司完成后返回首页", variable=self.return_home_var).grid(row=3, column=0, sticky="w", pady=6)

        ttk.Label(
            mode_card,
            text="这些选项主要用于调试。正常批量下载时，建议只保留“最后一家公司完成后返回首页”。",
            style="Subtle.TLabel",
            wraplength=340,
            justify="left",
        ).grid(row=4, column=0, sticky="w", pady=(12, 0))

        coordinate_card = ttk.LabelFrame(right, text="坐标校准", style="Card.TLabelframe")
        coordinate_card.grid(row=1, column=0, sticky="nsew", pady=(10, 0))
        coordinate_card.columnconfigure(0, weight=1)

        self._build_coordinate_panel(coordinate_card)

        actions = ttk.LabelFrame(right, text="执行", style="Card.TLabelframe")
        actions.grid(row=2, column=0, sticky="new", pady=(10, 0))
        actions.columnconfigure(0, weight=1)

        self.start_button = ttk.Button(actions, text="开始批量运行", command=self.start_run)
        self.start_button.configure(style="Accent.TButton")
        self.start_button.grid(row=0, column=0, sticky="ew")
        self.stop_button = ttk.Button(actions, text="停止运行", command=self.stop_run, state="disabled", style="Muted.TButton")
        self.stop_button.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        ttk.Button(actions, text="打开下载目录", command=self.open_output_dir, style="Muted.TButton").grid(row=2, column=0, sticky="ew", pady=(16, 0))
        ttk.Button(actions, text="打开日志目录", command=self.open_log_dir, style="Muted.TButton").grid(row=3, column=0, sticky="ew", pady=(8, 0))

        log_card = ttk.LabelFrame(paned, text="运行日志", style="Card.TLabelframe")
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(0, weight=1)

        self.log_text = tk.Text(
            log_card,
            wrap="word",
            height=10,
            bg="#0f1722",
            fg="#e8eef8",
            insertbackground="#e8eef8",
            relief="flat",
            padx=12,
            pady=12,
            font=("Consolas", 10),
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(log_card, orient="vertical", command=self.log_text.yview, style="App.Vertical.TScrollbar")
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.configure(cursor="xterm")

        paned.add(top_surface, weight=5)
        paned.add(log_card, weight=2)
        paned.sashpos(0, 690)

        status_bar = ttk.Frame(self.root, style="App.TFrame", padding=(20, 0, 20, 12))
        status_bar.grid(row=2, column=0, sticky="ew")
        status_bar.columnconfigure(0, weight=1)
        ttk.Label(status_bar, text="提示：运行前请确认 Choice 已登录；任务执行中尽量不要操作 Choice 窗口。", style="Subtle.TLabel").grid(row=0, column=0, sticky="w")

        self._bind_mousewheel(self.root)
        self._bind_mousewheel(self.content_canvas)
        self._bind_mousewheel(self.scrollable_body)
        self._bind_mousewheel(left)
        self._bind_mousewheel(right)
        self._bind_mousewheel(form)

    def _bind_mousewheel(self, widget):
        widget.bind("<MouseWheel>", self._on_mousewheel, add="+")
        widget.bind("<Button-4>", self._on_mousewheel, add="+")
        widget.bind("<Button-5>", self._on_mousewheel, add="+")

    def _on_mousewheel(self, event):
        if hasattr(event, "delta") and event.delta:
            self.content_canvas.yview_scroll(int(-event.delta / 120), "units")
        elif getattr(event, "num", None) == 4:
            self.content_canvas.yview_scroll(-1, "units")
        elif getattr(event, "num", None) == 5:
            self.content_canvas.yview_scroll(1, "units")
        return "break"

    def _on_scroll_body_configure(self, _event):
        self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.content_canvas.itemconfigure(self.content_canvas_window, width=event.width)

    def _add_path_row(
        self,
        parent,
        row: int,
        label: str,
        variable: tk.StringVar,
        command=None,
        button_text: str | None = None,
    ):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=8)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(10, 10), pady=8)
        if command and button_text:
            ttk.Button(parent, text=button_text, command=command).grid(row=row, column=2, sticky="ew", pady=8)

    def _build_coordinate_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        ttk.Label(
            parent,
            text="默认使用内置坐标。港股或不同分辨率下，可先测试鼠标位置，再填入新坐标并确认。",
            style="Subtle.TLabel",
            wraplength=620,
            justify="left",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        table = ttk.Frame(parent)
        table.grid(row=1, column=0, sticky="ew")
        table.columnconfigure(1, weight=1)
        ttk.Label(table, text="定位点").grid(row=0, column=0, sticky="w", pady=(0, 4))
        ttk.Label(table, text="X").grid(row=0, column=1, sticky="w", pady=(0, 4))
        ttk.Label(table, text="Y").grid(row=0, column=2, sticky="w", pady=(0, 4))
        ttk.Label(table, text="操作").grid(row=0, column=3, sticky="w", pady=(0, 4))

        for row_index, (target_name, label) in enumerate(COORDINATE_TARGET_LABELS, start=1):
            x_var, y_var = self.coordinate_vars[target_name]
            ttk.Label(table, text=label).grid(row=row_index, column=0, sticky="w", pady=3)
            ttk.Entry(table, textvariable=x_var, width=8).grid(row=row_index, column=1, sticky="w", padx=(8, 6), pady=3)
            ttk.Entry(table, textvariable=y_var, width=8).grid(row=row_index, column=2, sticky="w", padx=(0, 8), pady=3)
            row_actions = ttk.Frame(table)
            row_actions.grid(row=row_index, column=3, sticky="w", pady=3)
            ttk.Button(
                row_actions,
                text="确认",
                width=7,
                command=lambda name=target_name: self.confirm_coordinate(name),
            ).grid(row=0, column=0, sticky="w")
            ttk.Button(
                row_actions,
                text="取鼠标",
                width=8,
                command=lambda name=target_name: self.capture_mouse_coordinate(name),
            ).grid(row=0, column=1, sticky="w", padx=(6, 0))

        test_bar = ttk.Frame(parent)
        test_bar.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        test_bar.columnconfigure(1, weight=1)
        ttk.Label(test_bar, text="测试定位点").grid(row=0, column=0, sticky="w")
        ttk.Combobox(
            test_bar,
            textvariable=self.calibration_target_var,
            values=[label for _name, label in COORDINATE_TARGET_LABELS],
            state="readonly",
            width=18,
        ).grid(row=0, column=1, sticky="w", padx=(8, 8))
        self.calibration_button = ttk.Button(test_bar, text="执行到此处并移动鼠标", command=self.start_coordinate_calibration)
        self.calibration_button.grid(row=0, column=2, sticky="e")

    def _choose_excel_file(self):
        selected = filedialog.askopenfilename(
            parent=self.root,
            title="选择 Excel 文件",
            initialdir=str(APP_DIR),
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if selected:
            self.excel_path_var.set(selected)

    def _choose_choice_exe(self):
        selected = filedialog.askopenfilename(
            parent=self.root,
            title="选择 Choice 程序或快捷方式",
            initialdir=str(DEFAULT_BROWSE_DIR),
            filetypes=[("可执行文件或快捷方式", "*.exe *.lnk"), ("所有文件", "*.*")],
        )
        if selected:
            self.choice_exe_var.set(selected)

    def _choose_output_dir(self):
        self._choose_directory("选择下载目录", self.output_dir_var)

    def _choose_log_dir(self):
        self._choose_directory("选择日志目录", self.log_dir_var)

    def _choose_directory(self, title: str, variable: tk.StringVar):
        if self.is_running:
            messagebox.showinfo("任务进行中", "任务运行时不能修改目录。")
            return
        current_text = variable.get().strip()
        current_path = Path(current_text) if current_text else APP_DIR
        initial_dir = current_path if current_path.exists() else APP_DIR
        try:
            selected = self._show_windows_folder_dialog(title, initial_dir)
        except Exception as exc:
            self._write_gui_error(exc)
            messagebox.showerror("选择目录失败", f"无法打开目录选择窗口：{exc}")
            return
        if selected:
            variable.set(selected)

    def _show_windows_folder_dialog(self, title: str, initial_dir: Path) -> str:
        escaped_title = self._powershell_quote(title)
        escaped_initial_dir = self._powershell_quote(str(initial_dir))
        script = f"""
Add-Type -AssemblyName System.Windows.Forms
[System.Windows.Forms.Application]::EnableVisualStyles()
$dialog = New-Object System.Windows.Forms.FolderBrowserDialog
$dialog.Description = {escaped_title}
$dialog.SelectedPath = {escaped_initial_dir}
$dialog.ShowNewFolderButton = $true
$result = $dialog.ShowDialog()
if ($result -eq [System.Windows.Forms.DialogResult]::OK) {{
    [Console]::OutputEncoding = [System.Text.Encoding]::UTF8
    Write-Output $dialog.SelectedPath
}}
"""
        flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-STA",
                "-ExecutionPolicy",
                "Bypass",
                "-Command",
                script,
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=flags,
        )
        if result.returncode != 0:
            raise RuntimeError((result.stderr or result.stdout or "PowerShell folder dialog failed.").strip())
        return result.stdout.strip()

    def _powershell_quote(self, value: str) -> str:
        return "'" + value.replace("'", "''") + "'"

    def _write_gui_error(self, exc: Exception):
        try:
            ERROR_LOG_PATH.write_text(
                f"{exc}\n\n{traceback.format_exc()}",
                encoding="utf-8",
            )
        except Exception:
            pass

    def _target_name_from_label(self, label: str) -> str:
        for target_name, target_label in COORDINATE_TARGET_LABELS:
            if label == target_label:
                return target_name
        return COORDINATE_TARGET_LABELS[0][0]

    def _collect_coordinate_overrides(self) -> dict[str, tuple[int, int]]:
        overrides: dict[str, tuple[int, int]] = {}
        for target_name, _label in COORDINATE_TARGET_LABELS:
            x_var, y_var = self.coordinate_vars[target_name]
            raw_x = x_var.get().strip()
            raw_y = y_var.get().strip()
            if not raw_x or not raw_y:
                raise ValueError(f"请填写坐标：{target_name}")
            try:
                point = (int(raw_x), int(raw_y))
            except ValueError as exc:
                raise ValueError(f"坐标必须是整数：{target_name}") from exc
            if point != self.confirmed_coordinates[target_name]:
                label = COORDINATE_LABEL_BY_NAME.get(target_name, target_name)
                raise ValueError(f"{label} 坐标已修改，请点击该行“确认”后再运行或测试。")
            default_point = choice_automation.COORDINATE_DEFAULTS[target_name]
            if point != default_point:
                overrides[target_name] = point
        return overrides

    def _build_coordinate_args(self, coordinate_overrides: dict[str, tuple[int, int]]) -> list[str]:
        args: list[str] = []
        for name, point in coordinate_overrides.items():
            args.extend(["--coordinate", f"{name}={point[0]},{point[1]}"])
        return args

    def confirm_coordinate(self, target_name: str):
        try:
            x_var, y_var = self.coordinate_vars[target_name]
            x_value = int(x_var.get().strip())
            y_value = int(y_var.get().strip())
        except Exception:
            messagebox.showerror("坐标错误", "坐标必须填写为整数。")
            return
        self.confirmed_coordinates[target_name] = (x_value, y_value)
        label = COORDINATE_LABEL_BY_NAME.get(target_name, target_name)
        self._append_log(f"[GUI] 已确认坐标 {label}=({x_value}, {y_value})，后续运行将按当前坐标执行\n")

    def capture_mouse_coordinate(self, target_name: str):
        if self.is_running or self.is_calibrating:
            messagebox.showinfo("任务进行中", "任务运行或校准时不能读取鼠标坐标。")
            return
        choice_exe = self.choice_exe_var.get().strip()
        if not choice_exe:
            messagebox.showerror("参数错误", "请先填写 Choice 程序路径，用于连接 Choice 窗口。")
            return
        self._append_log(f"[GUI] 3 秒内请把鼠标移动到目标位置：{target_name}\n")
        self.root.after(3000, lambda: self._capture_mouse_coordinate_now(target_name, choice_exe))

    def _capture_mouse_coordinate_now(self, target_name: str, choice_exe: str):
        if self.is_running or self.is_calibrating:
            return
        try:
            window = choice_automation.launch_or_connect(Path(choice_exe))
            rect = window.rectangle()
            cursor_point = ctypes.wintypes.POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(cursor_point))
            relative_x = cursor_point.x - rect.left
            relative_y = cursor_point.y - rect.top
            x_var, y_var = self.coordinate_vars[target_name]
            x_var.set(str(relative_x))
            y_var.set(str(relative_y))
            self._append_log(f"[GUI] 已读取鼠标坐标 {target_name}=({relative_x}, {relative_y})\n")
        except Exception as exc:
            self._write_gui_error(exc)
            messagebox.showerror("读取坐标失败", str(exc))

    def _load_companies_from_excel(self, excel_path: Path) -> list[str]:
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("当前仅支持 .xlsx 或 .xlsm 文件。")

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                raise ValueError("Excel 文件为空。")

            header_map = {}
            for index, value in enumerate(header):
                if value is None:
                    continue
                name = str(value).strip()
                if name:
                    header_map[name] = index

            if "公司名称" not in header_map:
                raise ValueError("Excel 中未找到“公司名称”列。")

            company_index = header_map["公司名称"]
            companies: list[str] = []
            for row in rows:
                if company_index >= len(row):
                    continue
                cell_value = row[company_index]
                if cell_value is None:
                    continue
                company_name = str(cell_value).strip()
                if company_name:
                    companies.append(company_name)

            if not companies:
                raise ValueError("Excel 的“公司名称”列没有可用数据。")
            return companies
        finally:
            workbook.close()

    def _refresh_company_count_text(self):
        excel_path = self.excel_path_var.get().strip()
        if not excel_path:
            self.company_count_var.set("未选择 Excel 文件")
            self.range_start_var.set("")
            self.range_end_var.set("")
            return

        try:
            companies = self._load_companies_from_excel(Path(excel_path))
        except Exception as exc:
            self.company_count_var.set(f"文件校验未通过: {exc}")
            self.range_start_var.set("")
            self.range_end_var.set("")
            return

        total = len(companies)
        try:
            current_start = max(1, int(self.range_start_var.get() or 1))
        except Exception:
            current_start = 1
        try:
            current_end = max(1, int(self.range_end_var.get() or total))
        except Exception:
            current_end = total
        self.range_start_var.set(str(min(current_start, total)))
        self.range_end_var.set(str(min(max(current_end, int(self.range_start_var.get())), total)))
        self.company_count_var.set(f"已加载 {len(companies)} 家企业")

    def _on_excel_path_changed(self, *_args):
        self._refresh_company_count_text()

    def _append_log(self, text: str):
        self.log_text.insert("end", text)
        self.log_text.see("end")

    def _drain_output_queue(self):
        while True:
            try:
                item = self.output_queue.get_nowait()
            except queue.Empty:
                break
            if isinstance(item, tuple) and len(item) == 2 and item[0] == "status":
                self.status_var.set(item[1])
            else:
                self._append_log(str(item))
        self.root.after(150, self._drain_output_queue)

    def _collect_run_config(self):
        excel_path = self.excel_path_var.get().strip()
        choice_exe = self.choice_exe_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        log_dir = self.log_dir_var.get().strip()

        if not excel_path:
            raise ValueError("请选择 Excel 文件。")
        if not choice_exe:
            raise ValueError("请输入 Choice 程序路径。")
        if not output_dir:
            raise ValueError("请输入下载目录。")
        if not log_dir:
            raise ValueError("请输入日志目录。")

        if not self.batch_count_var.get().strip():
            raise ValueError("请填写每家公司下载篇数。")
        if not self.max_batch_count_var.get().strip():
            raise ValueError("请填写本次下载篇数上限。")
        batch_count = int(self.batch_count_var.get())
        max_batch_count = int(self.max_batch_count_var.get())
        if max_batch_count < 1:
            raise ValueError("下载篇数上限必须大于 0。")
        if batch_count < 1 or batch_count > max_batch_count:
            raise ValueError(f"每家公司下载篇数必须在 1 到 {max_batch_count} 之间，当前为 {batch_count}。")

        companies = self._load_companies_from_excel(Path(excel_path))
        total_companies = len(companies)
        if not self.range_start_var.get().strip() or not self.range_end_var.get().strip():
            raise ValueError("请填写提取范围。选择 Excel 后程序会自动填入默认范围，也可以手动修改。")
        range_start = int(self.range_start_var.get())
        range_end = int(self.range_end_var.get())
        if range_start < 1 or range_end < 1:
            raise ValueError("提取范围必须从第 1 家开始填写。")
        if range_start > range_end:
            raise ValueError("开始序号不能大于结束序号。")
        if range_start > total_companies or range_end > total_companies:
            raise ValueError(f"提取范围超出企业总数，当前 Excel 共 {total_companies} 家企业。")
        selected_companies = companies[range_start - 1 : range_end]
        if len(selected_companies) > 1 and self.skip_navigation_var.get():
            raise ValueError("批量运行时不能勾选“跳过导航”，否则无法切换到下一家公司。")
        coordinate_overrides = self._collect_coordinate_overrides()

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        Path(log_dir).mkdir(parents=True, exist_ok=True)

        return {
            "excel_path": excel_path,
            "choice_exe": choice_exe,
            "output_dir": output_dir,
            "log_dir": log_dir,
            "companies": selected_companies,
            "range_start": range_start,
            "range_end": range_end,
            "total_companies": total_companies,
            "batch_count": batch_count,
            "max_batch_count": max_batch_count,
            "filename_keywords": self.filename_keywords_var.get().strip(),
            "keyword_match_mode": self._normalize_keyword_match_mode(self.keyword_match_mode_var.get()),
            "latest_only": self.latest_only_var.get(),
            "start_wait": self.start_wait_var.get().strip(),
            "enter_wait": self.enter_wait_var.get().strip(),
            "post_f9_wait": self.post_f9_wait_var.get().strip(),
            "navigation_only": self.navigation_only_var.get(),
            "skip_navigation": self.skip_navigation_var.get(),
            "skip_folder_dialog": self.skip_folder_dialog_var.get(),
            "return_home": self.return_home_var.get(),
            "coordinate_overrides": coordinate_overrides,
        }

    def _build_args_for_company(self, config: dict, company_name: str, should_return_home: bool):
        output_dir = Path(config["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)

        argv = [
            "--exe",
            config["choice_exe"],
            "--report-name",
            company_name,
            "--allowed-report",
            company_name,
            "--base-dir",
            str(output_dir),
            "--log-dir",
            config["log_dir"],
            "--batch-count",
            str(config["batch_count"]),
            "--max-batch-count",
            str(config["max_batch_count"]),
            "--filename-keywords",
            config["filename_keywords"],
            "--keyword-match-mode",
            config["keyword_match_mode"],
            "--enter-wait-seconds",
            config["enter_wait"],
            "--post-f9-wait-seconds",
            config["post_f9_wait"],
        ]
        argv.extend(self._build_coordinate_args(config["coordinate_overrides"]))
        if config["navigation_only"]:
            argv.append("--navigation-only")
        if config["skip_navigation"]:
            argv.append("--skip-navigation")
        if config["skip_folder_dialog"]:
            argv.append("--skip-folder-dialog")
        if not config["latest_only"]:
            argv.append("--keep-all-matches")
        if not should_return_home:
            argv.append("--skip-return-home")
        return choice_automation.parse_args(argv)

    def _normalize_keyword_match_mode(self, value: str) -> str:
        normalized = (value or "").strip()
        if normalized in KEYWORD_MATCH_MODE_LABELS:
            return KEYWORD_MATCH_MODE_LABELS[normalized]
        lowered = normalized.lower()
        if lowered in {"and", "or"}:
            return lowered
        return "or"

    def _write_summary_workbook(self, config: dict, results: list[dict]) -> Path:
        output_dir = Path(config["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)
        summary_path = output_dir / "choice_batch_summary.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量结果"
        sheet.append(["序号", "公司名称", "状态", "退出码", "说明"])
        for result in results:
            sheet.append(
                [
                    result["index"],
                    result["company"],
                    result["status"],
                    result["code"],
                    result["message"],
                ]
            )
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 70
        workbook.save(summary_path)
        workbook.close()
        return summary_path

    def start_run(self):
        if self.is_running or self.is_calibrating:
            messagebox.showinfo("任务进行中", "当前已有任务在运行。")
            return

        try:
            if not self.start_wait_var.get().strip():
                raise ValueError("请填写启动等待秒数。")
            if not self.enter_wait_var.get().strip():
                raise ValueError("请填写 Enter 等待秒数。")
            if not self.post_f9_wait_var.get().strip():
                raise ValueError("请填写 F9 等待秒数。")
            float(self.start_wait_var.get().strip())
            float(self.enter_wait_var.get().strip())
            float(self.post_f9_wait_var.get().strip())
            config = self._collect_run_config()
        except Exception as exc:
            messagebox.showerror("参数错误", str(exc))
            return

        self.log_text.delete("1.0", "end")
        total = len(config["companies"])
        self._append_log(
            f"[GUI] 已选择第 {config['range_start']} 到第 {config['range_end']} 家，共 {total} 家企业，开始执行批量任务\n\n"
        )
        self.is_running = True
        self.stop_requested = False
        self.stop_event.clear()
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.status_var.set(f"运行中 (0/{total})")

        def worker():
            overall_code = 0
            companies = config["companies"]
            total_companies = len(companies)
            completed = 0
            results: list[dict] = []
            start_wait = max(0.0, float(config["start_wait"] or 0))

            if start_wait:
                self.output_queue.put(f"[GUI] 启动等待 {start_wait:g} 秒后开始执行\n")
                if self.stop_event.wait(start_wait):
                    self.output_queue.put("[GUI] 已收到停止请求，批量任务停止\n")
                    results.append(
                        {
                            "index": 0,
                            "company": "",
                            "status": "已停止",
                            "code": 2,
                            "message": "启动等待期间停止",
                        }
                    )
                    summary_path = self._write_summary_workbook(config, results)
                    self.output_queue.put(f"[GUI] 批量结果汇总: {summary_path}\n")
                    self.root.after(0, lambda: self._finish_run(2, 0, total_companies))
                    return

            for index, company_name in enumerate(companies, start=1):
                if self.stop_event.is_set():
                    self.output_queue.put("[GUI] 已收到停止请求，批量任务停止\n")
                    overall_code = 2
                    break
                should_return_home = index < total_companies or config["return_home"]
                args = self._build_args_for_company(config, company_name, should_return_home)
                self.output_queue.put(("status", f"运行中 ({index}/{total_companies})"))
                self.output_queue.put(f"[GUI] [{index}/{total_companies}] 开始处理: {company_name}\n")
                self.output_queue.put(f"[GUI] 下载目录: {Path(config['output_dir'])}\n")

                try:
                    code = choice_automation.run_with_args(
                        args,
                        extra_log_handlers=[self.log_handler],
                        stop_event=self.stop_event,
                    )
                    message = "完成" if code == 0 else f"执行失败，退出码: {code}"
                except Exception as exc:
                    code = 1
                    message = f"未捕获异常: {exc}"
                    self.output_queue.put(f"[GUI] [{company_name}] {message}\n")

                if code != 0:
                    overall_code = code
                    results.append(
                        {
                            "index": index,
                            "company": company_name,
                            "status": "已停止" if self.stop_event.is_set() else "失败",
                            "code": code,
                            "message": message,
                        }
                    )
                    if self.stop_event.is_set():
                        self.output_queue.put(f"[GUI] [{company_name}] 当前公司已停止\n")
                        break
                    else:
                        self.output_queue.put(f"[GUI] [{company_name}] 执行失败，继续处理下一家公司\n\n")
                        continue

                completed = index
                results.append(
                    {
                        "index": index,
                        "company": company_name,
                        "status": "成功",
                        "code": code,
                        "message": message,
                    }
                )
                self.output_queue.put(f"[GUI] [{index}/{total_companies}] 已完成: {company_name}\n\n")

            summary_path = self._write_summary_workbook(config, results)
            self.output_queue.put(f"[GUI] 批量结果汇总: {summary_path}\n")
            self.root.after(0, lambda: self._finish_run(overall_code, completed, total_companies))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def _collect_calibration_config(self) -> dict:
        if not self.start_wait_var.get().strip():
            raise ValueError("请填写启动等待秒数。")
        if not self.enter_wait_var.get().strip():
            raise ValueError("请填写 Enter 等待秒数。")
        if not self.post_f9_wait_var.get().strip():
            raise ValueError("请填写 F9 等待秒数。")
        float(self.start_wait_var.get().strip())
        float(self.enter_wait_var.get().strip())
        float(self.post_f9_wait_var.get().strip())

        excel_path = self.excel_path_var.get().strip()
        choice_exe = self.choice_exe_var.get().strip()
        log_dir = self.log_dir_var.get().strip()
        if not excel_path:
            raise ValueError("请选择 Excel 文件。")
        if not choice_exe:
            raise ValueError("请输入 Choice 程序路径。")
        if not log_dir:
            raise ValueError("请输入日志目录。")

        companies = self._load_companies_from_excel(Path(excel_path))
        if not companies:
            raise ValueError("Excel 中没有可用公司。")
        try:
            range_start = max(1, int(self.range_start_var.get() or 1))
        except Exception:
            range_start = 1
        if range_start > len(companies):
            raise ValueError(f"提取范围超出企业总数，当前 Excel 共 {len(companies)} 家企业。")

        Path(log_dir).mkdir(parents=True, exist_ok=True)
        calibration_target = self._target_name_from_label(self.calibration_target_var.get())
        return {
            "choice_exe": choice_exe,
            "log_dir": log_dir,
            "company": companies[range_start - 1],
            "start_wait": self.start_wait_var.get().strip(),
            "enter_wait": self.enter_wait_var.get().strip(),
            "post_f9_wait": self.post_f9_wait_var.get().strip(),
            "calibration_target": calibration_target,
            "coordinate_overrides": self._collect_coordinate_overrides(),
        }

    def start_coordinate_calibration(self):
        if self.is_running or self.is_calibrating:
            messagebox.showinfo("任务进行中", "当前已有任务在运行。")
            return
        try:
            config = self._collect_calibration_config()
        except Exception as exc:
            messagebox.showerror("参数错误", str(exc))
            return

        self.log_text.delete("1.0", "end")
        self._append_log(
            f"[GUI] 坐标测试开始：{config['calibration_target']}，使用公司：{config['company']}\n"
            "[GUI] 程序会执行到该步骤并移动鼠标，随后停止，不会继续下载。\n\n"
        )
        self.is_calibrating = True
        self.stop_requested = False
        self.stop_event.clear()
        self.start_button.configure(state="disabled")
        self.calibration_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.status_var.set("坐标测试中")

        def worker():
            start_wait = max(0.0, float(config["start_wait"] or 0))
            if start_wait:
                self.output_queue.put(f"[GUI] 启动等待 {start_wait:g} 秒后开始坐标测试\n")
                if self.stop_event.wait(start_wait):
                    self.root.after(0, lambda: self._finish_calibration(2))
                    return

            argv = [
                "--exe",
                config["choice_exe"],
                "--report-name",
                config["company"],
                "--allowed-report",
                config["company"],
                "--log-dir",
                config["log_dir"],
                "--batch-count",
                "1",
                "--max-batch-count",
                "1",
                "--enter-wait-seconds",
                config["enter_wait"],
                "--post-f9-wait-seconds",
                config["post_f9_wait"],
                "--calibrate-target",
                config["calibration_target"],
                "--skip-return-home",
            ]
            argv.extend(self._build_coordinate_args(config["coordinate_overrides"]))
            try:
                args = choice_automation.parse_args(argv)
                code = choice_automation.run_with_args(
                    args,
                    extra_log_handlers=[self.log_handler],
                    stop_event=self.stop_event,
                )
            except Exception as exc:
                code = 1
                self.output_queue.put(f"[GUI] 坐标测试异常: {exc}\n")
            self.root.after(0, lambda: self._finish_calibration(code))

        self.calibration_thread = threading.Thread(target=worker, daemon=True)
        self.calibration_thread.start()

    def _finish_calibration(self, code: int):
        self.is_calibrating = False
        self.start_button.configure(state="normal")
        self.calibration_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        if code == 0:
            self.status_var.set("坐标测试完成")
            self._append_log("[GUI] 坐标测试完成。请观察鼠标位置，如不准确可用“取鼠标”或手动修改坐标。\n")
        elif self.stop_requested or code == 2:
            self.status_var.set("坐标测试已停止")
            self._append_log("[GUI] 坐标测试已按请求停止\n")
        else:
            self.status_var.set("坐标测试失败")
            self._append_log(f"[GUI] 坐标测试失败，退出码: {code}\n")

    def _finish_run(self, code: int, completed: int, total: int):
        self.is_running = False
        self.start_button.configure(state="normal")
        self.calibration_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        if code == 0:
            self.status_var.set(f"运行完成 ({total}/{total})")
            self._append_log("[GUI] 批量任务执行完成\n")
            self._clear_runtime_cache()
        elif self.stop_requested or code == 2:
            self.status_var.set(f"已停止 ({completed}/{total})")
            self._append_log("[GUI] 批量任务已按请求停止\n")
            self._clear_runtime_cache()
        else:
            self.status_var.set(f"运行完成，有失败 ({completed}/{total})")
            self._append_log(f"[GUI] 批量任务已结束，存在失败企业，最后失败退出码: {code}\n")
            self._clear_runtime_cache()

    def stop_run(self):
        if not self.is_running and not self.is_calibrating:
            return
        self.stop_requested = True
        self.stop_event.set()
        self.stop_button.configure(state="disabled")
        self.status_var.set("停止中...")
        self._append_log("[GUI] 已请求停止，程序会在当前步骤结束后尽快停止\n")

    def open_output_dir(self):
        self._open_path(self.output_dir_var.get().strip())

    def open_log_dir(self):
        self._open_path(self.log_dir_var.get().strip())

    def _open_path(self, path_text: str):
        if not path_text:
            messagebox.showerror("路径为空", "请先输入或选择目录。")
            return
        try:
            path = Path(path_text)
            path.mkdir(parents=True, exist_ok=True)
            subprocess.Popen(["explorer", str(path)])
        except Exception as exc:
            messagebox.showerror("打开目录失败", f"无法打开目录：{exc}")

    def _load_settings(self):
        try:
            if SETTINGS_PATH.exists():
                SETTINGS_PATH.unlink()
        except Exception:
            pass

    def _save_settings(self):
        return

    def _clear_runtime_cache(self):
        try:
            if SETTINGS_PATH.exists():
                SETTINGS_PATH.unlink()
                self._append_log("[GUI] 已清理本地缓存设置\n")
        except Exception as exc:
            self._append_log(f"[GUI] 清理本地缓存设置失败: {exc}\n")

    def on_close(self):
        if self.is_running or self.is_calibrating:
            if not messagebox.askyesno("退出确认", "任务仍在运行，确定要关闭窗口吗？"):
                return
        self._save_settings()
        self.root.destroy()


def main():
    root = tk.Tk()
    AutomationGui(root)
    root.mainloop()


if __name__ == "__main__":
    main()
